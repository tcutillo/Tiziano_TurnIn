#TWS
from ibapi.client import EClient
from ibapi.wrapper import EWrapper
import time
import datetime
#myTools
from myTools.createContract import initContract
from myTools.fillData import fillData
from myTools.getDays import getDuration
from myTools.marketData import MarketData
from myTools.styleCells import styleMarketData
# for Excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

listOfMarketData = []


class TWSApp(EClient, EWrapper):
    def __init__(self, contract, nbDays, workbook):
        EClient.__init__(self, self)
        self.myContract = contract
        self.nbDays = nbDays
        self.workBook = workbook
        self.workSheet = workbook.create_sheet()
        self.workSheet.title = f"{nbDays}_Days"
        self.workSheetName = f"{nbDays}_Days"
        self.queryTime = (datetime.datetime.today()).strftime("%Y%m%d-%H:%M:%S")
        self.NTerm = 0
        self.totalVolumeYest = 0
        self.avgVolume = 0

    def nextValidId(self, orderId: int):
        self.workSheet.append(['N-term', 'date', 'time', 'yestCls', 'currentPrice', 'changePrevPrice$', 'changePrevPrice%', 'changeYestCls', 'Volume', 'yestPriorVolume', 'TotalYesterdayVolume', 'Avg Volume', 'Bid', 'Ask', 'BidAskSpread', 'minBid', 'maxAsk', 'maxBidAskSpread', 'Avg BidAskSpread', 'Minimum tolerance', 'Maximum tolerance'])
        self.reqMarketDataType(4)
        self.reqHistoricalData(orderId, self.myContract, self.queryTime, getDuration(self.nbDays), "10 mins", "TRADES", 1, 1, False, [])
        time.sleep(5)
        self.reqContractDetails(orderId, self.myContract)
        

    def contractDetails(self, reqId, contractDetails):
        self.workBook.save(f"{contractDetails.contract.symbol}.xlsx")
        self.workBookName = f"{contractDetails.contract.symbol}.xlsx"
    
    def contractDetailsEnd(self, reqId):
        self.disconnect()

    def historicalData(self, reqId, bar):
        dateSplit = bar.date.split()
        currentSecurity = MarketData()

        if self.NTerm < 39:
            currentSecurity.time = dateSplit[1]
            currentSecurity.currentPrice = bar.close
            currentSecurity.volume = bar.volume
            listOfMarketData.append(currentSecurity)
        
        elif self.NTerm >= 39:
            currentSecurity = fillData(self, currentSecurity, listOfMarketData, bar, dateSplit)
            if self.NTerm == 39:
                self.avgVolume = currentSecurity.volume
            else:
                self.avgVolume = (self.avgVolume + currentSecurity.volume)
            self.workSheet.append([currentSecurity.Nterm, currentSecurity.date, currentSecurity.time, currentSecurity.yestCls, currentSecurity.currentPrice, currentSecurity.changePrevPriceDollar, currentSecurity.changePrevPricePercentage, currentSecurity.changeYestCls, currentSecurity.volume, currentSecurity.yestPriorVolume, currentSecurity.totalVolumeYest])
            listOfMarketData.append(currentSecurity)
        self.NTerm += 1

def getMarketData(contract):
    workbook = Workbook()
    workbook.remove(workbook['Sheet'])
    listOfDays = [30, 60, 90]
    
    for nbDays in listOfDays:
        currentSecurity = TWSApp(contract, nbDays, workbook)
        currentSecurity.connect("127.0.0.1", 7497, 1)
        currentSecurity.run()
        sh = workbook[f"{currentSecurity.workSheetName}"]
        styleMarketData(sh, currentSecurity)
        # styleCells(sh, getAvgVolumeBar(currentSecurity.avgVolume, nbDays), getAvgBidAskSpread(currentSecurityBidAsk.avgBidAskSpread, nbDays))
        workbook.save(currentSecurity.workBookName)
        listOfMarketData = []
        
def main():
    # listContracts = initContract([296574745, 297249704, 7089, 470458975, 4215217, 13977, 6890, 5684], ["NYMEX", "IPE", "NYSE", "NYSE", "NYSE", "NYSE", "NYSE", "NYSE"]) #CL1, WTI, FCX, CHK, XLE, XOM, EOG, CVX
    # listContracts = initContract([296574745], ["NYMEX"])  #CL1        80$
    # listContracts = initContract([7089], ["NYSE"])        #FCX        45$
    # listContracts = initContract([470458975], ["NYSE"])   #CHK        90$  
    # listContracts = initContract([4215217], ["NYSE"])     #XLE        87$
    listContracts = initContract([13977], ["NYSE"])         #XOM       110$
    # listContracts = initContract([6890], ["NYSE"])        #EOG       129$
    # listContracts = initContract([5684], ["NYSE"])        #CVX       177$

    for contract in listContracts:
        getMarketData(contract)
    return 0

if __name__ == "__main__":
    main()