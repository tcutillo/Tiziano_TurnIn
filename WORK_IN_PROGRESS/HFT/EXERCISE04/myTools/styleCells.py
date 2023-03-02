from decimal import Decimal
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

from myTools.getDays import getAvgBidAskSpread, getAvgVolumeBar



def changeLengthCells(sh):
    sh.column_dimensions['A'].width = 6
    sh.column_dimensions['B'].width = 9
    sh.column_dimensions['C'].width = 8.5
    sh.column_dimensions['D'].width = 8
    sh.column_dimensions['E'].width = 11.5
    sh.column_dimensions['F'].width = 15.5
    sh.column_dimensions['G'].width = 15.5
    sh.column_dimensions['H'].width = 13
    sh.column_dimensions['I'].width = 9
    sh.column_dimensions['J'].width = 14.5
    sh.column_dimensions['K'].width = 18.5
    sh.column_dimensions['L'].width = 13
    sh.column_dimensions['M'].width = 13
    sh.column_dimensions['N'].width = 13
    sh.column_dimensions['O'].width = 13
    sh.column_dimensions['P'].width = 13
    sh.column_dimensions['Q'].width = 13
    sh.column_dimensions['R'].width = 16
    sh.column_dimensions['S'].width = 18
    sh.column_dimensions['T'].width = 18
    sh.column_dimensions['U'].width = 18

def addBidAskToSh(sh, currentSecurity, row):

    sh['M' + str(row)].value = currentSecurity.Bid
    sh['N' + str(row)].value = currentSecurity.Ask
    sh['O' + str(row)].value = currentSecurity.BidAskSpread
    sh['P' + str(row)].value = currentSecurity.minBid
    sh['Q' + str(row)].value = currentSecurity.maxAsk
    sh['R' + str(row)].value = currentSecurity.maxBidAskSpread

def styleMarketData(sh, currentSecurity):
    changeLengthCells(sh)
    for row in range(2, sh.max_row+1):
        sh['L' + str(row)].value = round(getAvgVolumeBar(currentSecurity.avgVolume, currentSecurity.nbDays))


def styleBidAsk(sh, currentSecurity):
    i = round(getAvgBidAskSpread(currentSecurity.avgBidAskSpread, currentSecurity.nbDays), 3)
    win = 0
    for row in range(2, sh.max_row+1):
        sh['S' + str(row)].value = i
        ii = sh['R' + str(row)].value
        iii = (ii - i) / 2
        iv = abs(i - iii)
        v = iii + i
        sh['T' + str(row)].value = iv
        sh['U' + str(row)].value = v
        bidAskSpread = sh['O' + str(row)].value
        if sh['C' + str(row)].value == "09:30:00":
                win += sh['H' + str(row)].value
        if sh['I' + str(row)].value > (sh['L' + str(row)].value * Decimal(2.5)) and bidAskSpread < v:
            for column in range(1, sh.max_column+1):
                sh[get_column_letter(column) + str(row)].fill = PatternFill("solid", start_color="00008000")
        if sh['I' + str(row)].value > (sh['L' + str(row)].value / Decimal(2.5)) and bidAskSpread > v:
            for column in range(1, sh.max_column+1):
                sh[get_column_letter(column) + str(row)].fill = PatternFill("solid", start_color="00FF0000")
    print(f"win per share: {win}\n")